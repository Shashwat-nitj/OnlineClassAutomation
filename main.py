import time

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from win10toast import ToastNotifier

GOOGLE = "https://www.google.com"

opt = Options()
opt.add_argument("--disable-infobars")
opt.add_argument("start-maximized")
opt.add_argument("--disable-extensions")
opt.add_experimental_option("prefs", {
    "profile.default_content_setting_values.media_stream_mic": 2,
    "profile.default_content_setting_values.media_stream_camera": 2,
    "profile.default_content_setting_values.geolocation": 2,
    "profile.default_content_setting_values.notifications": 2
})


def now(): return int(datetime.now().strftime('%H%M'))


def now1(): return '[' + str(datetime.now().strftime('%H:%M:%S')) + ']'


def print1(s): ToastNotifier().show_toast("Online Class Automation", s, duration=2)


def click(xpath): driver.find_element(By.XPATH, xpath).click()


def sendkeys(xpath, key): driver.find_element(By.XPATH, xpath).send_keys(key)


def td(t): return datetime.strptime(t + '00', '%H%M%S') - datetime.now()


def at_time(t):
    if t < 1000:
        time_diff = td('0' + str(t))
    else:
        time_diff = td(str(t))
    time.sleep(time_diff.seconds)


def join_meet():
    print(now1(), "---Initializing Google Meet---")
    try:
        driver.get(link)
        click('//*[@id="yDmH0d"]/div[3]/div/div[2]/div[3]/div/span/span')  # Dismiss (without audio video)
        time.sleep(3)
        click(
            '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[3]/div/div/div[2]/div/div[2]/div/div[1]/div['  # Join
            '1]/span/span')
        print(now1(), "Joining...")
        print1("Meeting Joined")

        at_time(end)
        print(now1(), "Leaving...")
        print1("Leaving Meeting")
        click('//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[2]/div/div[7]/span/button/i')  # Leave
        driver.get(GOOGLE)
    except Exception as e:
        print("Error Initializing Meet")
        driver.get(GOOGLE)


def join_team():
    print(now1(), "---Initializing Microsoft Teams---")
    try:
        driver.get(link)
        click('//*[@id="buttonsbox"]/button[2]')  # Use Web Version
        driver.implicitly_wait(10)
        click(
            '//*[@id="ngdialog1"]/div[2]/div/div/div/div[1]/div/div/div[2]/div/button')  # Continue without audio or video
        click('//*[@id="page-content-wrapper"]/div[1]/div/calling-pre-join-screen/div/div/div[2]/div[1]/div['  # Join
              '2]/div/div/section/div[1]/div/div/button')
        print(now1(), "Joining...")
        print1("Meeting Joined")
        at_time(end)
        print(now1(), "Leaving...")
        print1("Leaving Meeting")
        driver.get(GOOGLE)
    except Exception as e:
        print("Error Initializing Teams")
        driver.get(GOOGLE)


def join_class():
    if link[8] == 'm':
        if sign_in['google']:
            join_meet()
        else:
            print1("Meet Link Provided but google account not signed in")
            print("Check Google Account Credentials")
    else:
        if sign_in['microsoft']:
            join_team()
        else:
            print1("Teams Link Provided but microsoft account not signed in")
            print("Check Microsoft Account Credentials")


def google_sign_in(email, password):
    try:
        click('//*[@id="gb"]/div/div[2]/a')  # Sign in
        sendkeys('//*[@id="identifierId"]', email)  # Email
        driver.implicitly_wait(10)
        click('//*[@id="identifierNext"]/div/button/span')  # Next
        sendkeys('//*[@id="password"]/div[1]/div/div[1]/input', password)  # Password
        click('//*[@id="passwordNext"]/div/button/span')  # Next
        sendkeys('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input', "Signed In!!!")
        print1("Google Sign In Success")
        time.sleep(1)
        sign_in['google'] = True
    except Exception as e:
        print1("Error Signing In Google")
        driver.get(GOOGLE)


def microsoft_sign_in(email, password):
    try:
        driver.get('https://www.microsoft.com/en-in/microsoft-teams/log-in')
        click(
            '//*[@id="office-HeroPhotographic-6j70rzy"]/div/div[3]/section/div/div[2]/div[1]/a')  # Sign In Btn (Homepage)
        driver.switch_to.window(driver.window_handles[1])  # Switching to signIn page
        driver.implicitly_wait(10)
        sendkeys('//*[@id="i0116"]', email)  # Email
        click('//*[@id="idSIButton9"]')  # Next
        sendkeys('//*[@id="i0118"]', password)  # Password
        time.sleep(1)
        click('//*[@id="idSIButton9"]')  # Submit
        time.sleep(0.1)
        click(
            '//*[@id="lightbox"]/div[3]/div/div[2]/div/div[3]/div[1]/div/label/span')  # Don't show message again(check)
        click('//*[@id="idSIButton9"]')  # Keep me logged in (Yes)
        time.sleep(2)
        driver.switch_to.window(driver.window_handles[0])
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        print1("Microsoft Sign In Success")
        time.sleep(1)
        sign_in['microsoft'] = True
    except Exception as e:
        print1("Error Signing In Microsoft")
        driver.get(GOOGLE)


def get_info():
    wb = load_workbook('automation.xlsx')
    wb1 = wb['Sheet 1']
    link_info = []
    c = datetime.today().weekday() * 3 + 1
    r = 3
    while type(wb1.cell(r, c + 1).value) == str:
        link_info.append([wb1.cell(r, c).value, wb1.cell(r, c + 1).value, wb1.cell(r, c + 2).value])
        r += 1
    return [link_info, wb1.cell(15, 1).value, wb1.cell(15, 2).value, wb1.cell(17, 1).value, wb1.cell(17, 2).value]


# Day's Beginning
driver = webdriver.Chrome(chrome_options=opt)
driver.get(GOOGLE)
driver.maximize_window()

print(now1(), "Hello World!")
info = get_info()
g_email, g_pass = info[1], info[2]
ms_email, ms_pass = info[3], info[4]
sign_in = {'microsoft': False, 'google': True}
if type(g_email) == str and type(g_pass) == str:
    print(now1(), "---Google---")
    google_sign_in(g_email, g_pass)
else:
    print("---Google Account Info NOT GOOD---")
if type(ms_email) == str and type(ms_pass) == str:
    print(now1(), "---Microsoft---")
    microsoft_sign_in(ms_email, ms_pass)
else:
    print("---Microsoft Account Info NOT GOOD---")

# Sorting classes info list on the basis of start time
info[0].sort()
for i in info[0]:
    print('')
    start, link, end = i[0], i[1], i[0] + i[2]
    print("Current Time:", now1())
    print("Next Class Time:", start)
    if end < now():  # Missed classes
        print("---Class missed---")
        continue
    else:
        if start > now():
            print1('Next Class in ' + str(td(str(start)).seconds) + ' seconds')
            at_time(start)  # Waiting for class to start
        join_class()  # Join the class, and to end it.

print1("Class List End")
print('\n', now1(), "---SHUTTING driver DOWN---")
print1('Closing Driver...')
driver.close()
